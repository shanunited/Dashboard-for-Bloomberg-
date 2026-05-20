from __future__ import annotations

import re
from datetime import date
from io import BytesIO
from typing import Any

import pandas as pd
from fastapi import FastAPI, File, HTTPException, Request, UploadFile
from fastapi.exceptions import RequestValidationError
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
from openpyxl import load_workbook

try:
    from .report_generator import generate_combined_pdf, generate_pdf
except ImportError:
    from report_generator import generate_combined_pdf, generate_pdf


app = FastAPI(title="ULJK Bloomberg Ranking API", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.exception_handler(RequestValidationError)
async def validation_exception_handler(request: Request, exc: RequestValidationError) -> JSONResponse:
    if request.url.path == "/index-rp":
        for error in exc.errors():
            if "index_rp_file" in error.get("loc", []):
                return JSONResponse(
                    status_code=400,
                    content={"detail": "Index RP Excel file is required."},
                )
    return JSONResponse(status_code=422, content={"detail": exc.errors()})


REQUIRED_ALIAS_MAP = {
    "company_name": ["Company Name"],
    "index_clean": ["Index_Clean"],
    "cmp": ["CMP"],
    "fs": ["FS", "FUNDAMENTAL SCORE", "Fundamental Score"],
    "ts": ["TS", "TECHNICAL SCORE", "Technical Score"],
    "total": ["Total"],
    "rsi": ["RSI"],
    "wma_30": ["30WMA"],
    "wma_30_pct": ["% Dist 150 DMA"],
    "roce": ["ROCE FY26", "ROCE FY25", "ROCE 2026", "ROCE"],
    "roe": ["ROE FY26", "ROE FY25", "ROE 2026", "ROE"],
    "eps_growth": ["EPS Growth", "EPS 5Y CAGR"],
    "d_e": ["D/E", "D/E Ratio"],
}

NUMERIC_COLUMNS = [
    "cmp",
    "fs",
    "ts",
    "total",
    "rsi",
    "wma_30",
    "wma_30_pct",
    "roce",
    "roe",
    "eps_growth",
    "d_e",
]

INDEX_RP_ALIAS_MAP = {
    "index_name": ["Index", "Index Name", "Sector", "Name"],
    "ltp": ["LTP", "Rate", "Last Price", "CMP"],
    "daily_rp_call": ["Daily RP Call", "Daily RP Calls"],
    "weekly_rp_call": ["Weekly RP Call", "Weekly RP Calls"],
    "rsi": ["RSI", "RSI 14D"],
    "wma_30": ["30WMA", "30 WMA", "30WMA Value"],
    "rsi_score": ["RSI SCORE", "RSI Score"],
    "wma_30_score": ["30WMA SCORE", "30WMA Score", "30 WMA Score"],
    "weekly_rp_score": ["MRS WEEKLY SCORE", "Weekly RP Score"],
    "daily_rp_score": ["MRS DAILY SCORE", "Daily RP Score"],
    "total": ["TOTAL", "Total", "Score"],
}

INDEX_RP_REQUIRED_FIELDS = [
    "index_name",
    "ltp",
    "daily_rp_call",
    "weekly_rp_call",
    "rsi",
    "wma_30",
    "rsi_score",
    "wma_30_score",
    "weekly_rp_score",
    "daily_rp_score",
    "total",
]

INDEX_RP_NUMERIC_FIELDS = [
    "ltp",
    "rsi",
    "rsi_score",
    "wma_30_score",
    "weekly_rp_score",
    "daily_rp_score",
    "total",
]


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def normalize_optional_text(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    return normalize_text(value)


def canonical_key(value: Any) -> str:
    text = normalize_text(value).lower()
    return "".join(character for character in text if character.isalnum())


def deduplicate_headers(headers: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    resolved: list[str] = []
    for header in headers:
        count = seen.get(header, 0)
        seen[header] = count + 1
        resolved.append(header if count == 0 else f"{header}.{count}")
    return resolved


def find_header_row(worksheet) -> int | None:
    for row_index, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=25, values_only=True),
        start=1,
    ):
        values = {canonical_key(cell) for cell in row if normalize_text(cell)}
        if "companyname" in values and "indexclean" in values:
            return row_index
    return None


def build_headers(worksheet, header_row_index: int) -> list[str]:
    header_row = list(
        next(
            worksheet.iter_rows(
                min_row=header_row_index,
                max_row=header_row_index,
                values_only=True,
            )
        )
    )
    group_row: list[Any] | None = None
    if header_row_index > 1:
        group_row = list(
            next(
                worksheet.iter_rows(
                    min_row=header_row_index - 1,
                    max_row=header_row_index - 1,
                    values_only=True,
                )
            )
        )

    headers: list[str] = []
    active_group = ""
    for column_index, header_value in enumerate(header_row):
        group_value = group_row[column_index] if group_row and column_index < len(group_row) else None
        group_text = normalize_text(group_value)
        header_text = normalize_text(header_value)

        if group_text:
            active_group = group_text

        if header_text.startswith("FY") and active_group:
            headers.append(f"{active_group} {header_text}")
        elif header_text:
            headers.append(header_text)
        elif group_text:
            headers.append(group_text)
        else:
            headers.append(f"Unnamed_{column_index}")

    return deduplicate_headers(headers)


def read_workbook(file_bytes: bytes) -> pd.DataFrame:
    workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=False)
    target_sheet_name: str | None = None
    header_row_index: int | None = None
    headers: list[str] | None = None

    for worksheet in workbook.worksheets:
        found_row = find_header_row(worksheet)
        if found_row is not None:
            target_sheet_name = worksheet.title
            header_row_index = found_row
            headers = build_headers(worksheet, found_row)
            break

    workbook.close()

    if target_sheet_name is None or header_row_index is None or headers is None:
        raise ValueError("Could not find a sheet/header row containing Company Name and Index_Clean.")

    dataframe = pd.read_excel(
        BytesIO(file_bytes),
        sheet_name=target_sheet_name,
        header=None,
        skiprows=header_row_index,
        names=headers,
    )
    return dataframe.dropna(how="all").reset_index(drop=True)


def resolve_columns(dataframe: pd.DataFrame) -> dict[str, str]:
    normalized_lookup = {
        normalize_text(column).lower(): column for column in dataframe.columns
    }
    canonical_lookup: dict[str, list[str]] = {}
    for column in dataframe.columns:
        canonical_lookup.setdefault(canonical_key(column), []).append(column)
    resolved: dict[str, str] = {}

    for output_name, aliases in REQUIRED_ALIAS_MAP.items():
        match = None
        for alias in aliases:
            alias_key = normalize_text(alias).lower()
            if alias_key in normalized_lookup:
                match = normalized_lookup[alias_key]
                break
        if match is None:
            for alias in aliases:
                alias_key = canonical_key(alias)
                if alias_key in canonical_lookup:
                    match = canonical_lookup[alias_key][0]
                    break
        if match is None:
            raise ValueError(f"Required column not found for {output_name}: {aliases}")
        resolved[output_name] = match

    return resolved


def resolve_alias_columns(
    dataframe: pd.DataFrame,
    alias_map: dict[str, list[str]],
    required_fields: list[str],
) -> dict[str, str]:
    normalized_lookup = {
        normalize_text(column).lower(): column for column in dataframe.columns
    }
    canonical_lookup: dict[str, list[str]] = {}
    for column in dataframe.columns:
        canonical_lookup.setdefault(canonical_key(column), []).append(column)

    resolved: dict[str, str] = {}
    missing: list[str] = []

    for output_name, aliases in alias_map.items():
        match = None
        for alias in aliases:
            alias_key = normalize_text(alias).lower()
            if alias_key in normalized_lookup:
                match = normalized_lookup[alias_key]
                break
        if match is None:
            for alias in aliases:
                alias_key = canonical_key(alias)
                if alias_key in canonical_lookup:
                    match = canonical_lookup[alias_key][0]
                    break
        if match is None:
            if output_name in required_fields:
                missing.append(f"{output_name}: {', '.join(aliases)}")
        else:
            resolved[output_name] = match

    if missing:
        raise ValueError("Required columns missing from first sheet: " + "; ".join(missing))

    return resolved


def read_index_rp_sheet(file_bytes: bytes) -> pd.DataFrame:
    try:
        workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("Invalid Excel format. Please upload a valid Excel workbook.") from exc

    try:
        sheet_names = workbook.sheetnames
    finally:
        workbook.close()

    if not sheet_names:
        raise ValueError("Workbook has no sheets.")

    try:
        dataframe = pd.read_excel(BytesIO(file_bytes), sheet_name=0)
    except Exception as exc:
        raise ValueError("Invalid Excel format. Please upload a valid Excel workbook.") from exc

    dataframe = dataframe.dropna(how="all").reset_index(drop=True)
    if dataframe.empty:
        raise ValueError("First sheet is empty.")

    dataframe.columns = [normalize_text(column) for column in dataframe.columns]
    if all(str(column).startswith("Unnamed") for column in dataframe.columns):
        raise ValueError("Required columns missing from first sheet.")

    return dataframe


def clean_index_rp_dataframe(file_bytes: bytes) -> pd.DataFrame:
    raw = read_index_rp_sheet(file_bytes)
    column_map = resolve_alias_columns(raw, INDEX_RP_ALIAS_MAP, INDEX_RP_REQUIRED_FIELDS)

    cleaned = pd.DataFrame(
        {
            "index_name": raw[column_map["index_name"]],
            "ltp": raw[column_map["ltp"]],
            "daily_rp_call": raw[column_map["daily_rp_call"]],
            "weekly_rp_call": raw[column_map["weekly_rp_call"]],
            "rsi": raw[column_map["rsi"]],
            "wma_30": raw[column_map["wma_30"]],
            "rsi_score": raw[column_map["rsi_score"]],
            "wma_30_score": raw[column_map["wma_30_score"]],
            "weekly_rp_score": raw[column_map["weekly_rp_score"]],
            "daily_rp_score": raw[column_map["daily_rp_score"]],
            "total": raw[column_map["total"]],
        }
    )

    cleaned["index_name"] = cleaned["index_name"].map(normalize_optional_text)
    cleaned["daily_rp_call"] = cleaned["daily_rp_call"].map(lambda value: normalize_optional_text(value).upper())
    cleaned["weekly_rp_call"] = cleaned["weekly_rp_call"].map(lambda value: normalize_optional_text(value).upper())
    cleaned["wma_30"] = cleaned["wma_30"].map(lambda value: normalize_optional_text(value).title())

    for column in INDEX_RP_NUMERIC_FIELDS:
        cleaned[column] = pd.to_numeric(cleaned[column], errors="coerce")

    cleaned = cleaned.dropna(subset=["index_name", "total"]).copy()
    cleaned = cleaned[cleaned["index_name"] != ""].copy()
    if cleaned.empty:
        raise ValueError("First sheet does not contain any valid index rows.")

    cleaned = cleaned.sort_values(
        by=["total", "index_name"],
        ascending=[False, True],
        kind="stable",
    ).reset_index(drop=True)

    return cleaned


def build_index_rp_payload(dataframe: pd.DataFrame) -> dict[str, Any]:
    strongest = dataframe.iloc[0]["index_name"]
    weakest = dataframe.sort_values(
        by=["total", "index_name"],
        ascending=[True, True],
        kind="stable",
    ).iloc[0]["index_name"]

    kpis = {
        "total_indexes": int(len(dataframe)),
        "daily_buy_calls": int(dataframe["daily_rp_call"].eq("BUY").sum()),
        "weekly_buy_calls": int(dataframe["weekly_rp_call"].eq("BUY").sum()),
        "above_30wma": int(dataframe["wma_30"].str.upper().eq("ABOVE").sum()),
        "strongest_index": make_json_value(strongest),
        "weakest_index": make_json_value(weakest),
    }

    rows = []
    for _, row in dataframe.iterrows():
        rows.append(
            {
                "index_name": make_json_value(row["index_name"]),
                "ltp": make_json_value(row["ltp"]),
                "daily_rp_call": make_json_value(row["daily_rp_call"]),
                "weekly_rp_call": make_json_value(row["weekly_rp_call"]),
                "rsi": make_json_value(row["rsi"]),
                "wma_30": make_json_value(row["wma_30"]),
                "rsi_score": make_json_value(row["rsi_score"]),
                "wma_30_score": make_json_value(row["wma_30_score"]),
                "weekly_rp_score": make_json_value(row["weekly_rp_score"]),
                "daily_rp_score": make_json_value(row["daily_rp_score"]),
                "total": make_json_value(row["total"]),
            }
        )

    return {"kpis": kpis, "rows": rows}


def clean_dataframe(file_bytes: bytes) -> pd.DataFrame:
    raw = read_workbook(file_bytes)
    column_map = resolve_columns(raw)

    cleaned = pd.DataFrame(
        {
            "company_name": raw[column_map["company_name"]],
            "index_clean": raw[column_map["index_clean"]],
            "cmp": raw[column_map["cmp"]],
            "fs": raw[column_map["fs"]],
            "ts": raw[column_map["ts"]],
            "total": raw[column_map["total"]],
            "rsi": raw[column_map["rsi"]],
            "wma_30": raw[column_map["wma_30"]],
            "wma_30_pct": raw[column_map["wma_30_pct"]],
            "roce": raw[column_map["roce"]],
            "roe": raw[column_map["roe"]],
            "eps_growth": raw[column_map["eps_growth"]],
            "d_e": raw[column_map["d_e"]],
        }
    )

    cleaned["company_name"] = cleaned["company_name"].map(normalize_text)
    cleaned["index_clean"] = cleaned["index_clean"].map(normalize_text)

    for column in NUMERIC_COLUMNS:
        cleaned[column] = pd.to_numeric(cleaned[column], errors="coerce")

    cleaned = cleaned.dropna(subset=["company_name", "index_clean", "total"]).copy()
    cleaned = cleaned[
        (cleaned["company_name"] != "") & (cleaned["index_clean"] != "")
    ].copy()

    cleaned = cleaned.sort_values(
        by=["index_clean", "total", "company_name"],
        ascending=[True, False, True],
        kind="stable",
    )
    cleaned = cleaned.drop_duplicates(subset=["index_clean", "company_name"], keep="first")
    cleaned = cleaned.reset_index(drop=True)
    return cleaned


def add_rankings(dataframe: pd.DataFrame) -> pd.DataFrame:
    top_sorted = dataframe.sort_values(
        by=["index_clean", "total", "company_name"],
        ascending=[True, False, True],
        kind="stable",
    ).copy()
    top_sorted["top_rank"] = top_sorted.groupby("index_clean").cumcount() + 1

    bottom_sorted = dataframe.sort_values(
        by=["index_clean", "total", "company_name"],
        ascending=[True, True, True],
        kind="stable",
    ).copy()
    bottom_sorted["bottom_rank"] = bottom_sorted.groupby("index_clean").cumcount() + 1

    ranked = dataframe.merge(
        top_sorted[["index_clean", "company_name", "top_rank"]],
        on=["index_clean", "company_name"],
        how="left",
    )
    ranked = ranked.merge(
        bottom_sorted[["index_clean", "company_name", "bottom_rank"]],
        on=["index_clean", "company_name"],
        how="left",
    )
    return ranked


def build_change(today_rank: Any, yesterday_rank: Any, *, mode: str) -> str:
    if pd.isna(yesterday_rank):
        return "New Entry"
    if pd.isna(today_rank):
        return "Missing Today"

    delta = int(yesterday_rank - today_rank)
    if mode == "top":
        if delta > 0:
            return f"Up by {delta}"
        if delta < 0:
            return f"Down by {abs(delta)}"
        return "Unchanged"

    if today_rank < yesterday_rank:
        return f"Down by {int(yesterday_rank - today_rank)}"
    if today_rank > yesterday_rank:
        return f"Up by {int(today_rank - yesterday_rank)}"
    return "Unchanged"


def make_json_value(value: Any) -> Any:
    if pd.isna(value):
        return None
    if isinstance(value, (pd.Timestamp,)):
        return value.isoformat()
    if hasattr(value, "item"):
        value = value.item()
    return value


def extract_report_date(filename: str | None) -> str:
    if filename:
        match = re.search(r"(\d{1,2})[-_](\d{1,2})[-_](\d{2,4})", filename)
        if match:
            day, month, year = match.groups()
            if len(year) == 2:
                year = f"20{year}"
            return f"{int(day):02d}-{int(month):02d}-{year}"
    return date.today().strftime("%d-%m-%Y")


def build_index_payload(today_ranked: pd.DataFrame, yesterday_ranked: pd.DataFrame) -> dict[str, Any]:
    merged = today_ranked.merge(
        yesterday_ranked[
            ["index_clean", "company_name", "top_rank", "bottom_rank", "total"]
        ].rename(
            columns={
                "top_rank": "yesterday_top_rank",
                "bottom_rank": "yesterday_bottom_rank",
                "total": "yesterday_total",
            }
        ),
        on=["index_clean", "company_name"],
        how="left",
    )

    merged["score_change"] = merged["total"] - merged["yesterday_total"]
    merged["top_change"] = merged.apply(
        lambda row: build_change(row["top_rank"], row["yesterday_top_rank"], mode="top"),
        axis=1,
    )
    merged["bottom_change"] = merged.apply(
        lambda row: build_change(
            row["bottom_rank"],
            row["yesterday_bottom_rank"],
            mode="bottom",
        ),
        axis=1,
    )

    response: dict[str, Any] = {"indices": [], "data": {}}

    for index_name, index_frame in merged.groupby("index_clean", sort=True):
        top10 = index_frame.sort_values(
            by=["top_rank", "company_name"],
            ascending=[True, True],
            kind="stable",
        ).head(10)
        bottom10 = index_frame.sort_values(
            by=["bottom_rank", "company_name"],
            ascending=[True, True],
            kind="stable",
        ).head(10)

        summary = {
            "stocks_in_index": int(len(index_frame)),
            "new_top10_entries": int(
                ((top10["yesterday_top_rank"].isna()) | (top10["yesterday_top_rank"] > 10)).sum()
            ),
            "new_bottom10_entries": int(
                (
                    (bottom10["yesterday_bottom_rank"].isna())
                    | (bottom10["yesterday_bottom_rank"] > 10)
                ).sum()
            ),
            "top10_unchanged": int(top10["top_change"].eq("Unchanged").sum()),
            "bottom10_unchanged": int(bottom10["bottom_change"].eq("Unchanged").sum()),
        }

        top10_rows = []
        for _, row in top10.iterrows():
            top10_rows.append(
                {
                    "top_rank": make_json_value(row["top_rank"]),
                    "company_name": make_json_value(row["company_name"]),
                    "today_total": make_json_value(row["total"]),
                    "yesterday_top_rank": make_json_value(row["yesterday_top_rank"]),
                    "change": make_json_value(row["top_change"]),
                    "cmp": make_json_value(row["cmp"]),
                    "fs": make_json_value(row["fs"]),
                    "ts": make_json_value(row["ts"]),
                    "rsi": make_json_value(row["rsi"]),
                    "30wma": make_json_value(row["wma_30"]),
                    "30wma_pct": make_json_value(row["wma_30_pct"]),
                    "roce": make_json_value(row["roce"]),
                    "roe": make_json_value(row["roe"]),
                    "eps_growth": make_json_value(row["eps_growth"]),
                    "d_e": make_json_value(row["d_e"]),
                    "score_change": make_json_value(row["score_change"]),
                }
            )

        bottom10_rows = []
        for _, row in bottom10.iterrows():
            bottom10_rows.append(
                {
                    "bottom_rank": make_json_value(row["bottom_rank"]),
                    "company_name": make_json_value(row["company_name"]),
                    "today_total": make_json_value(row["total"]),
                    "yesterday_bottom_rank": make_json_value(row["yesterday_bottom_rank"]),
                    "change": make_json_value(row["bottom_change"]),
                    "cmp": make_json_value(row["cmp"]),
                    "fs": make_json_value(row["fs"]),
                    "ts": make_json_value(row["ts"]),
                    "rsi": make_json_value(row["rsi"]),
                    "30wma": make_json_value(row["wma_30"]),
                    "30wma_pct": make_json_value(row["wma_30_pct"]),
                    "roce": make_json_value(row["roce"]),
                    "roe": make_json_value(row["roe"]),
                    "eps_growth": make_json_value(row["eps_growth"]),
                    "d_e": make_json_value(row["d_e"]),
                    "score_change": make_json_value(row["score_change"]),
                }
            )

        response["indices"].append(index_name)
        response["data"][index_name] = {
            "summary": summary,
            "top10": top10_rows,
            "bottom10": bottom10_rows,
        }

    return response


@app.get("/")
def root() -> dict[str, str]:
    return {"message": "ULJK Bloomberg Ranking API is running."}


@app.post("/generate-report")
async def generate_report(
    report_file: UploadFile = File(...),
    report_date: str = "Report",
) -> StreamingResponse:
    if not report_file.filename:
        raise HTTPException(status_code=400, detail="Excel file is required.")
    try:
        file_bytes = await report_file.read()
        pdf_bytes = generate_pdf(file_bytes, report_date)
        filename = f"Stock_Report_{report_date}.pdf"
        return StreamingResponse(
            iter([pdf_bytes]),
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Report generation failed: {exc}") from exc


@app.post("/generate-combined-report")
async def generate_combined_report(
    stock_file: UploadFile | None = File(None),
    index_file: UploadFile | None = File(None),
) -> StreamingResponse:
    if stock_file is None or not stock_file.filename:
        raise HTTPException(status_code=400, detail="Daily Stock Dashboard Sheet is required.")
    if index_file is None or not index_file.filename:
        raise HTTPException(status_code=400, detail="Daily Index Sheet is required.")

    try:
        stock_bytes = await stock_file.read()
        index_bytes = await index_file.read()
        report_date = extract_report_date(index_file.filename)
        pdf_bytes = generate_combined_pdf(stock_bytes, index_bytes, report_date)
        filename = f"Combined_Index_Stock_Report_{report_date}.pdf"
        return StreamingResponse(
            iter([pdf_bytes]),
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"PDF generation failed: {exc}") from exc


@app.post("/index-rp")
async def process_index_rp(index_rp_file: UploadFile | None = File(None)) -> dict[str, Any]:
    if index_rp_file is None or not index_rp_file.filename:
        raise HTTPException(status_code=400, detail="Index RP Excel file is required.")

    try:
        file_bytes = await index_rp_file.read()
        index_rp_clean = clean_index_rp_dataframe(file_bytes)
        return build_index_rp_payload(index_rp_clean)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:  # pragma: no cover - API safety
        raise HTTPException(status_code=500, detail=f"Index RP processing failed: {exc}") from exc


@app.post("/compare")
async def compare_files(
    yesterday_file: UploadFile = File(...),
    today_file: UploadFile = File(...),
) -> dict[str, Any]:
    if not yesterday_file.filename or not today_file.filename:
        raise HTTPException(status_code=400, detail="Both Excel files are required.")

    try:
        yesterday_bytes = await yesterday_file.read()
        today_bytes = await today_file.read()

        yesterday_clean = clean_dataframe(yesterday_bytes)
        today_clean = clean_dataframe(today_bytes)

        yesterday_ranked = add_rankings(yesterday_clean)
        today_ranked = add_rankings(today_clean)

        return build_index_payload(today_ranked, yesterday_ranked)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:  # pragma: no cover - API safety
        raise HTTPException(status_code=500, detail=f"Comparison failed: {exc}") from exc
