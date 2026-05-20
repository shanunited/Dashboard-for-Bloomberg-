"""Daily Stock Scoring PDF report generator."""
from __future__ import annotations

import io
from html import escape
from typing import Any

import pandas as pd
from openpyxl import load_workbook as xl_load
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (
    BaseDocTemplate,
    Frame,
    KeepTogether,
    PageBreak,
    PageTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_LEFT, TA_RIGHT

PAGE_W, PAGE_H = A4          # 595.28 x 841.89 pts
MARGIN = 14 * mm
FOOTER_H = 13 * mm
USABLE_W = PAGE_W - 2 * MARGIN

# ── Color palette ─────────────────────────────────────────────────────────────
CN    = colors.HexColor("#0a1628")   # deep navy (headers)
CN2   = colors.HexColor("#0e2038")   # lighter navy (sub-headers)
CGOLD = colors.HexColor("#f0a500")   # gold
CGND  = colors.HexColor("#0c2a10")   # green dark bg (top cards)
CGNA  = colors.HexColor("#27ae60")   # green accent
CRDD  = colors.HexColor("#2a0c0c")   # red dark bg (bottom cards)
CRDA  = colors.HexColor("#c0392b")   # red accent
CTXT  = colors.HexColor("#dbe7f2")   # main text
CMUT  = colors.HexColor("#8ea3b8")   # muted text
CBG   = colors.HexColor("#09111b")   # card body bg
CBDR  = colors.HexColor("#243447")   # border
CTICK = colors.HexColor("#4ade80")   # tick green
CCRSS = colors.HexColor("#f87171")   # cross red
CW    = colors.white


# ── Paragraph styles ──────────────────────────────────────────────────────────
def _ps(name: str, **kw) -> ParagraphStyle:
    p = ParagraphStyle(name)
    for k, v in kw.items():
        setattr(p, k, v)
    return p


ST_CVRTITLE = _ps("cvrtitle", fontName="Helvetica-Bold", fontSize=24, textColor=CW,    alignment=TA_CENTER, spaceAfter=6)
ST_CVRSUB   = _ps("cvrsub",   fontName="Helvetica",      fontSize=13, textColor=CGOLD, alignment=TA_CENTER, spaceAfter=4)
ST_IDXHDR   = _ps("idxhdr",   fontName="Helvetica-Bold", fontSize=13, textColor=CW,    alignment=TA_CENTER)
ST_SECHDR   = _ps("sechdr",   fontName="Helvetica-Bold", fontSize=9,  textColor=CGOLD, alignment=TA_LEFT)
ST_COMPANY  = _ps("company",  fontName="Helvetica-Bold", fontSize=10, textColor=CW,    alignment=TA_LEFT)
ST_CMP      = _ps("cmp_s",    fontName="Helvetica-Bold", fontSize=10, textColor=CGOLD, alignment=TA_RIGHT)
ST_SCLBL    = _ps("sclbl",    fontName="Helvetica",      fontSize=7,  textColor=CMUT,  alignment=TA_CENTER)
ST_SCPOS    = _ps("scpos",    fontName="Helvetica-Bold", fontSize=13, textColor=CTICK, alignment=TA_CENTER)
ST_SCNEG    = _ps("scneg",    fontName="Helvetica-Bold", fontSize=13, textColor=CCRSS, alignment=TA_CENTER)
ST_SCNEU    = _ps("scneu",    fontName="Helvetica-Bold", fontSize=13, textColor=CGOLD, alignment=TA_CENTER)
ST_INDHDR   = _ps("indhdr",   fontName="Helvetica-Bold", fontSize=8,  textColor=CMUT,  alignment=TA_LEFT)
ST_TICK     = _ps("tick_s",   fontName="Helvetica",      fontSize=8,  textColor=CTICK, alignment=TA_LEFT, leading=11)
ST_CROSS    = _ps("cross_s",  fontName="Helvetica",      fontSize=8,  textColor=CCRSS, alignment=TA_LEFT, leading=11)
ST_MVAL     = _ps("mval",     fontName="Helvetica",      fontSize=8,  textColor=CTXT,  alignment=TA_LEFT, leading=12)
ST_NARR     = _ps("narr",     fontName="Helvetica",      fontSize=8,  textColor=CTXT,  alignment=TA_JUSTIFY, leading=12)
ST_LEGEND   = _ps("legend",   fontName="Helvetica",      fontSize=9,  textColor=CTXT,  alignment=TA_LEFT, leading=13)
ST_LEGKEY   = _ps("legkey",   fontName="Helvetica-Bold", fontSize=9,  textColor=CGOLD, alignment=TA_LEFT, leading=13)
ST_COVBODY  = _ps("covbody",  fontName="Helvetica",      fontSize=10, textColor=CTXT,  alignment=TA_LEFT, leading=14)
ST_TBLHDR   = _ps("tblhdr",   fontName="Helvetica-Bold", fontSize=6.5, textColor=CGOLD, alignment=TA_CENTER, leading=8)
ST_TBLCELL  = _ps("tblcell",  fontName="Helvetica",      fontSize=6.3, textColor=CTXT,  alignment=TA_LEFT, leading=7.5)
ST_TBLNUM   = _ps("tblnum",   fontName="Helvetica",      fontSize=6.3, textColor=CTXT,  alignment=TA_RIGHT, leading=7.5)
ST_SMALLTITLE = _ps("smalltitle", fontName="Helvetica-Bold", fontSize=16, textColor=CW, alignment=TA_CENTER, spaceAfter=4)


# ── Column alias map for the report ──────────────────────────────────────────
REPORT_COLS: dict[str, list[str]] = {
    # Core
    "company_name":   ["Company Name"],
    "index_clean":    ["Index_Clean"],
    "cmp":            ["CMP"],
    "fs":             ["FS", "FUNDAMENTAL SCORE", "Fundamental Score"],
    "ts":             ["TS", "TECHNICAL SCORE", "Technical Score"],
    "total":          ["Total"],
    # Technical score indicators (binary)
    "rsir":           ["RSIR"],
    "wma30r":         ["30war"],
    "weekly_rs":      ["52 WR"],
    "daily_rs":       ["52 DR"],
    # Fundamental score indicators (binary)
    "sales_yoy":      ["SALES-R(YOY)"],
    "sales_qoq":      ["SALES-R(QOQ)"],
    "rocer":          ["ROCER"],
    "roer":           ["ROER"],
    "de_ratio":       ["D/E Ratio"],
    "fcf_r":          ["FCF-R"],
    "eps_s":          ["EPS(S)"],
    "roce5y_score":   ["ROCE5Y CAGR"],
    "pat_margin":     ["PAT Margin"],
    "ebitda_margin":  ["EBITA Margin"],
    # Raw metric values
    "rsi_val":        ["RSI.1"],
    "yoy_pct":        ["YOY%"],
    "qoq_pct":        ["QOQ%"],
    "de_raw":         ["D/E"],
    "ebitda_pct":     ["Ebitda%"],
    "roce5y_raw":     ["ROCE 5Y CAGR"],
    "eps5y_cagr":     ["EPS 5Y CAGR"],
}

FUND_INDICATORS = [
    ("sales_yoy",    "Sales Growth YoY"),
    ("sales_qoq",    "Sales Growth QoQ"),
    ("rocer",        "ROCE"),
    ("roer",         "ROE"),
    ("de_ratio",     "D/E Ratio"),
    ("fcf_r",        "FCF"),
    ("eps_s",        "EPS"),
    ("roce5y_score", "ROCE 5Y CAGR"),
    ("pat_margin",   "PAT Margin"),
    ("ebitda_margin","EBITDA Margin"),
]

TECH_INDICATORS = [
    ("rsir",      "RSI"),
    ("wma30r",    "30 WMA"),
    ("weekly_rs", "Weekly RS (52W)"),
    ("daily_rs",  "Daily RS (52W)"),
]

COMBINED_INDEX_COLS: dict[str, list[str]] = {
    "index_name": ["INDEX", "Index", "Index Name", "Sector", "Name"],
    "ltp": ["Rate", "LTP", "Last Price", "CMP"],
    "daily_rp_call": ["Daily RP Call", "Daily RP Calls"],
    "weekly_rp_call": ["Weekly RP Call", "Weekly RP Calls"],
    "rsi": ["RSI", "RSI 14D"],
    "wma_30": ["30WMA", "30 WMA", "30WMA Value"],
    "rsi_score": ["RSI SCORE", "RSI Score"],
    "wma_30_score": ["30WMA SCORE", "30WMA Score", "30 WMA Score"],
    "weekly_rp_score": ["MRS WEEKLY SCORE", "Weekly RP Score"],
    "daily_rp_score": ["MRS DAILY SCORE", "Daily RP Score"],
    "total": ["TOTAL", "Total", "Score"],
}

COMBINED_STOCK_COLS: dict[str, list[str]] = {
    "index_clean": ["Index_Clean", "index_clean", "Index Clean", "INDEX_CLEAN", "Index", "Sector"],
    "company_name": ["Company Name", "company name", "Company", "COMPANY NAME", "Stock Name"],
    "fs": ["FS", "Fundamental Score", "FUNDAMENTAL SCORE"],
    "ts": ["TS", "Technical Score", "TECHNICAL SCORE"],
    "total": ["Total", "TOTAL", "Final Score", "Final Total"],
}


# ── Excel reading utilities (self-contained, no import from main.py) ──────────
def _norm(v: Any) -> str:
    if v is None:
        return ""
    return " ".join(str(v).strip().split())


def _ckey(v: Any) -> str:
    t = _norm(v).lower()
    return "".join(c for c in t if c.isalnum())


def _dedup(headers: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    out: list[str] = []
    for h in headers:
        n = seen.get(h, 0)
        seen[h] = n + 1
        out.append(h if n == 0 else f"{h}.{n}")
    return out


def _find_header_row(ws) -> int | None:
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=25, values_only=True), start=1):
        vals = {_ckey(c) for c in row if _norm(c)}
        if "companyname" in vals and "indexclean" in vals:
            return i
    return None


def _build_headers(ws, hi: int) -> list[str]:
    hrow = list(next(ws.iter_rows(min_row=hi, max_row=hi, values_only=True)))
    grow: list[Any] | None = None
    if hi > 1:
        grow = list(next(ws.iter_rows(min_row=hi - 1, max_row=hi - 1, values_only=True)))

    headers: list[str] = []
    active_g = ""
    for ci, hv in enumerate(hrow):
        gv = grow[ci] if grow and ci < len(grow) else None
        gt = _norm(gv)
        ht = _norm(hv)
        if gt:
            active_g = gt
        if ht.startswith("FY") and active_g:
            headers.append(f"{active_g} {ht}")
        elif ht:
            headers.append(ht)
        elif gt:
            headers.append(gt)
        else:
            headers.append(f"Unnamed_{ci}")
    return _dedup(headers)


def _read_raw(file_bytes: bytes) -> pd.DataFrame:
    wb = xl_load(io.BytesIO(file_bytes), read_only=True, data_only=False)
    sheet_name = hi = headers = None
    for ws in wb.worksheets:
        found = _find_header_row(ws)
        if found is not None:
            sheet_name = ws.title
            hi = found
            headers = _build_headers(ws, found)
            break
    wb.close()
    if sheet_name is None:
        raise ValueError("Could not find a sheet containing Company Name and Index_Clean headers.")
    df = pd.read_excel(
        io.BytesIO(file_bytes),
        sheet_name=sheet_name,
        header=None,
        skiprows=hi,
        names=headers,
    )
    return df.dropna(how="all").reset_index(drop=True)


def _resolve(df: pd.DataFrame, aliases: list[str]) -> str | None:
    """Return the actual DataFrame column matching the first alias found, or None."""
    norm_lu = {_norm(c).lower(): c for c in df.columns}
    can_lu: dict[str, list[str]] = {}
    for c in df.columns:
        can_lu.setdefault(_ckey(c), []).append(c)

    for alias in aliases:
        key = _norm(alias).lower()
        if key in norm_lu:
            return norm_lu[key]
    for alias in aliases:
        key = _ckey(alias)
        if key in can_lu:
            return can_lu[key][0]
    return None


def read_report_dataframe(file_bytes: bytes) -> pd.DataFrame:
    """Read Excel and return a clean DataFrame with all report columns."""
    raw = _read_raw(file_bytes)

    out_cols: dict[str, Any] = {}
    for field, aliases in REPORT_COLS.items():
        col = _resolve(raw, aliases)
        out_cols[field] = raw[col] if col is not None else pd.Series([None] * len(raw))

    df = pd.DataFrame(out_cols)
    df["company_name"] = df["company_name"].map(lambda x: _norm(x) if x is not None else "")
    df["index_clean"]  = df["index_clean"].map(lambda x: _norm(x) if x is not None else "")

    numeric_fields = [f for f in REPORT_COLS if f not in ("company_name", "index_clean")]
    for f in numeric_fields:
        df[f] = pd.to_numeric(df[f], errors="coerce")

    df = df.dropna(subset=["company_name", "index_clean", "total"]).copy()
    df = df[(df["company_name"] != "") & (df["index_clean"] != "")].copy()

    # Add ranks
    df = df.sort_values(
        ["index_clean", "total", "company_name"], ascending=[True, False, True]
    ).reset_index(drop=True)
    df["top_rank"] = df.groupby("index_clean").cumcount() + 1

    df_bot = df.sort_values(
        ["index_clean", "total", "company_name"], ascending=[True, True, True]
    ).copy()
    df_bot["bot_rank"] = df_bot.groupby("index_clean").cumcount() + 1

    df = df.merge(
        df_bot[["index_clean", "company_name", "bot_rank"]],
        on=["index_clean", "company_name"],
        how="left",
    )
    return df


# ── Helpers ───────────────────────────────────────────────────────────────────
def _isna(val: Any) -> bool:
    if val is None:
        return True
    try:
        return bool(pd.isna(val))
    except Exception:
        return False


def _fmt(val: Any, decimals: int = 1, suffix: str = "", mult: float = 1.0) -> str:
    if _isna(val):
        return "—"
    try:
        return f"{float(val) * mult:.{decimals}f}{suffix}"
    except (TypeError, ValueError):
        return "—"


def _is_positive(val: Any) -> bool:
    if _isna(val):
        return False
    try:
        return float(val) > 0
    except (TypeError, ValueError):
        return False


def _tc(val: Any, label: str) -> Paragraph:
    """Return a tick or cross paragraph for an indicator."""
    pos = _is_positive(val)
    return Paragraph(f"{'✓' if pos else '✗'}  {label}", ST_TICK if pos else ST_CROSS)


def _score_cell(val: Any, label: str, col_w: float) -> Table:
    """Small 1×2 table: label row + coloured score row."""
    if not _isna(val):
        try:
            v = float(val)
            st = ST_SCPOS if v > 0 else ST_SCNEG
            vs = f"{v:.1f}"
        except (TypeError, ValueError):
            st = ST_SCNEU
            vs = "—"
    else:
        st = ST_SCNEU
        vs = "—"

    t = Table([[Paragraph(label, ST_SCLBL)], [Paragraph(vs, st)]], colWidths=[col_w])
    t.setStyle(TableStyle([
        ("TOPPADDING",    (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
    ]))
    return t


# ── Narrative ─────────────────────────────────────────────────────────────────
def _build_narrative(row: dict, rank: int, is_top: bool, index_name: str, total_in_index: int) -> str:
    name   = row.get("company_name") or "This stock"
    total  = _fmt(row.get("total"), 1)
    fs     = _fmt(row.get("fs"),    1)
    ts     = _fmt(row.get("ts"),    1)
    rsi    = _fmt(row.get("rsi_val"), 1)

    if is_top:
        s1 = (
            f"{name} ranks #{rank} out of {total_in_index} stocks in {index_name} "
            f"with a Total Score of {total} (FS: {fs}, TS: {ts}), placing it among the top performers."
        )
    else:
        s1 = (
            f"{name} ranks #{rank} from the bottom out of {total_in_index} stocks in {index_name} "
            f"with a Total Score of {total} (FS: {fs}, TS: {ts}), flagging it as a laggard."
        )

    pos_fund = [lbl for key, lbl in FUND_INDICATORS if _is_positive(row.get(key))]
    neg_fund = [lbl for key, lbl in FUND_INDICATORS if not _is_positive(row.get(key))]

    if is_top:
        s2 = (
            f"Key fundamental strengths: {', '.join(pos_fund[:4])}."
            if pos_fund
            else "Fundamental indicators are predominantly weak despite a strong technical score."
        )
    else:
        s2 = (
            f"Fundamental concerns include: {', '.join(neg_fund[:4])}."
            if neg_fund
            else "Fundamental indicators are mostly healthy, suggesting the weakness is technical."
        )

    pos_tech = [lbl for key, lbl in TECH_INDICATORS if _is_positive(row.get(key))]
    n_pos = len(pos_tech)
    qual = "strong" if n_pos >= 3 else ("moderate" if n_pos >= 2 else "weak")
    s3 = f"Technical setup is {qual} with {n_pos}/4 positive signals and RSI at {rsi}."

    return f"{s1}  {s2}  {s3}"


# ── Stock card ────────────────────────────────────────────────────────────────
def _stock_card(row: dict, rank: int, is_top: bool, index_name: str, total_in_index: int) -> KeepTogether:
    hdr_bg = CN
    body_bg = CGND if is_top else CRDD
    acc = CGNA if is_top else CRDA
    W = USABLE_W
    HW = W / 2
    TW = W / 3

    cmp_str = "₹" + _fmt(row.get("cmp"), 2) if not _isna(row.get("cmp")) else "—"
    tag = f"#{rank}  {'▲ TOP' if is_top else '▼ BOTTOM'}"

    # ── Header: Company name + CMP ────────────────────────────────────────────
    hdr = Table(
        [[Paragraph(f"<b>{tag}</b>   {row.get('company_name', '—')}", ST_COMPANY),
          Paragraph(f"CMP: {cmp_str}", ST_CMP)]],
        colWidths=[W * 0.72, W * 0.28],
    )
    hdr.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), hdr_bg),
        ("TOPPADDING",    (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ("LEFTPADDING",   (0, 0), (0,  -1), 8),
        ("RIGHTPADDING",  (-1, 0), (-1, -1), 8),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("LINEBELOW",     (0, 0), (-1, -1), 1.5, acc),
    ]))

    # ── Scores row ────────────────────────────────────────────────────────────
    scores = Table(
        [[_score_cell(row.get("total"), "TOTAL SCORE", TW - 2),
          _score_cell(row.get("fs"),    "FUNDAMENTAL",  TW - 2),
          _score_cell(row.get("ts"),    "TECHNICAL",    TW - 2)]],
        colWidths=[TW, TW, TW],
    )
    scores.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), body_bg),
        ("TOPPADDING",    (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("INNERGRID",     (0, 0), (-1, -1), 0.5, CBDR),
        ("VALIGN",        (0, 0), (-1, -1), "TOP"),
        ("ALIGN",         (0, 0), (-1, -1), "CENTER"),
    ]))

    # ── Scorecards: Fundamentals + Technicals ─────────────────────────────────
    def _ind_table(items: list, col_w: float) -> Table:
        data = [[Paragraph("", ST_INDHDR)]] + [[_tc(row.get(k), lbl)] for k, lbl in items]
        t = Table(data, colWidths=[col_w - 6])
        t.setStyle(TableStyle([
            ("TOPPADDING",    (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ("LEFTPADDING",   (0, 0), (-1, -1), 6),
        ]))
        return t

    fund_block = Table(
        [[Paragraph("FUNDAMENTALS", ST_INDHDR)], [_ind_table(FUND_INDICATORS, HW)]],
        colWidths=[HW - 1],
    )
    tech_block = Table(
        [[Paragraph("TECHNICALS", ST_INDHDR)], [_ind_table(TECH_INDICATORS, HW)]],
        colWidths=[HW - 1],
    )
    for blk in (fund_block, tech_block):
        blk.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (0, 0), CN2),
            ("BACKGROUND",    (0, 1), (-1, -1), CBG),
            ("TOPPADDING",    (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING",   (0, 0), (-1, -1), 6),
            ("LINEBELOW",     (0, 0), (0, 0), 0.5, acc),
        ]))

    scorecard_row = Table([[fund_block, tech_block]], colWidths=[HW, HW])
    scorecard_row.setStyle(TableStyle([
        ("VALIGN",    (0, 0), (-1, -1), "TOP"),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, CBDR),
        ("TOPPADDING",    (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ("LEFTPADDING",   (0, 0), (-1, -1), 0),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 0),
    ]))

    # ── Raw metrics row ───────────────────────────────────────────────────────
    rsi    = _fmt(row.get("rsi_val"),    1)
    yoy    = _fmt(row.get("yoy_pct"),    1, "%")
    qoq    = _fmt(row.get("qoq_pct"),    1, "%")
    de     = _fmt(row.get("de_raw"),     2)
    ebitda = _fmt(row.get("ebitda_pct"), 1, "%")
    roce5y = _fmt(row.get("roce5y_raw"), 1, "%", mult=100)
    eps5y  = _fmt(row.get("eps5y_cagr"), 1, "%")

    metrics_text = (
        f"<b>RSI:</b> {rsi}  ·  <b>Revenue YoY:</b> {yoy}  ·  <b>Revenue QoQ:</b> {qoq}  ·  "
        f"<b>D/E:</b> {de}  ·  <b>EBITDA:</b> {ebitda}  ·  "
        f"<b>ROCE 5Y CAGR:</b> {roce5y}  ·  <b>EPS 5Y CAGR:</b> {eps5y}"
    )
    met_row = Table([[Paragraph(metrics_text, ST_MVAL)]], colWidths=[W])
    met_row.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), CN2),
        ("TOPPADDING",    (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING",   (0, 0), (-1, -1), 8),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
    ]))

    # ── Narrative row ─────────────────────────────────────────────────────────
    narrative = _build_narrative(row, rank, is_top, index_name, total_in_index)
    narr_row = Table([[Paragraph(narrative, ST_NARR)]], colWidths=[W])
    narr_row.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), CBG),
        ("TOPPADDING",    (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING",   (0, 0), (-1, -1), 8),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
        ("LINEABOVE",     (0, 0), (-1, 0),  0.5, CBDR),
    ]))

    # ── Outer card wrapper ────────────────────────────────────────────────────
    card = Table(
        [[hdr], [scores], [scorecard_row], [met_row], [narr_row]],
        colWidths=[W],
    )
    card.setStyle(TableStyle([
        ("VALIGN",        (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING",    (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ("LEFTPADDING",   (0, 0), (-1, -1), 0),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 0),
        ("BOX",           (0, 0), (-1, -1), 1, acc),
    ]))

    return KeepTogether([card, Spacer(1, 4 * mm)])


# ── Page chrome ───────────────────────────────────────────────────────────────
def _add_footer(canvas, doc, date_str: str) -> None:
    canvas.saveState()
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(CMUT)
    canvas.drawString(MARGIN, 7 * mm, "Proprietary & Confidential")
    canvas.drawRightString(PAGE_W - MARGIN, 7 * mm, f"Page {doc.page}  |  {date_str}")
    canvas.setStrokeColor(CBDR)
    canvas.setLineWidth(0.5)
    canvas.line(MARGIN, 11 * mm, PAGE_W - MARGIN, 11 * mm)
    canvas.restoreState()


def _index_banner(index_name: str) -> Table:
    t = Table([[Paragraph(index_name.upper(), ST_IDXHDR)]], colWidths=[USABLE_W])
    t.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), CN),
        ("TOPPADDING",    (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("LINEBELOW",     (0, 0), (-1, -1), 2, CGOLD),
    ]))
    return t


def _section_banner(text: str, acc: colors.Color) -> Table:
    t = Table([[Paragraph(text, ST_SECHDR)]], colWidths=[USABLE_W])
    t.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), CN2),
        ("TOPPADDING",    (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING",   (0, 0), (-1, -1), 8),
        ("LINEABOVE",     (0, 0), (-1, 0),  1, acc),
        ("LINEBELOW",     (0, 0), (-1, -1), 0.5, CBDR),
    ]))
    return t


# ── Cover page ────────────────────────────────────────────────────────────────
def _read_first_sheet(file_bytes: bytes, label: str) -> pd.DataFrame:
    try:
        wb = xl_load(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Invalid {label} Excel file.") from exc
    try:
        sheet_names = wb.sheetnames
    finally:
        wb.close()
    if not sheet_names:
        raise ValueError(f"{label.capitalize()} workbook has no sheets.")
    try:
        df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=0)
    except Exception as exc:
        raise ValueError(f"Invalid {label} Excel file.") from exc
    df = df.dropna(how="all").reset_index(drop=True)
    if df.empty:
        raise ValueError(f"Empty {label} sheet.")
    df.columns = [_norm(c) for c in df.columns]
    return df


def _require_columns(df: pd.DataFrame, alias_map: dict[str, list[str]], label: str) -> dict[str, str]:
    resolved: dict[str, str] = {}
    missing: list[str] = []
    for field, aliases in alias_map.items():
        col = _resolve(df, aliases)
        if col is None:
            missing.append(f"{field}: {', '.join(aliases)}")
        else:
            resolved[field] = col
    if missing:
        raise ValueError(f"Missing required {label} columns: " + "; ".join(missing))
    return resolved


def _header_key(value: Any) -> str:
    return _norm(value).lower()


def _read_stock_sheet_with_detected_header(file_bytes: bytes) -> pd.DataFrame:
    try:
        wb = xl_load(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("Invalid stock Excel file.") from exc
    try:
        sheet_names = wb.sheetnames
    finally:
        wb.close()

    if not sheet_names:
        raise ValueError("Stock workbook has no sheets.")

    try:
        raw = pd.read_excel(io.BytesIO(file_bytes), sheet_name=0, header=None)
    except Exception as exc:
        raise ValueError("Invalid stock Excel file.") from exc

    raw = raw.dropna(how="all").reset_index(drop=True)
    if raw.empty:
        raise ValueError("Empty stock sheet.")

    required_groups = [
        {"company name", "company", "stock name"},
        {"index_clean", "index clean", "index", "sector"},
        {"total", "final score", "final total"},
    ]
    preferred_groups = [
        {"fs", "fundamental score"},
        {"ts", "technical score"},
    ]
    best_row: int | None = None
    best_score = -1

    for row_index in range(min(10, len(raw))):
        values = {_header_key(cell) for cell in raw.iloc[row_index].tolist() if _norm(cell)}
        if all(values & group for group in required_groups):
            score = sum(1 for group in preferred_groups if values & group)
            if score > best_score:
                best_row = row_index
                best_score = score
            if score == len(preferred_groups):
                break

    if best_row is None:
        raise ValueError(
            "Could not detect stock file header row. Expected columns like Company Name, Index_Clean, FS, TS, and Total."
        )

    headers = [_norm(value) or f"Unnamed_{i}" for i, value in enumerate(raw.iloc[best_row].tolist())]
    data = raw.iloc[best_row + 1:].copy()
    data.columns = _dedup(headers)
    data = data.dropna(how="all").reset_index(drop=True)
    if data.empty:
        raise ValueError("Stock sheet does not contain any rows after the detected header.")
    return data


def _clean_call(value: Any) -> str:
    return _norm(value).upper()


def read_combined_index_dataframe(file_bytes: bytes) -> pd.DataFrame:
    raw = _read_first_sheet(file_bytes, "index")
    columns = _require_columns(raw, COMBINED_INDEX_COLS, "index")
    df = pd.DataFrame({
        "index_name": raw[columns["index_name"]],
        "ltp": raw[columns["ltp"]],
        "daily_rp_call": raw[columns["daily_rp_call"]],
        "weekly_rp_call": raw[columns["weekly_rp_call"]],
        "rsi": raw[columns["rsi"]],
        "wma_30": raw[columns["wma_30"]],
        "rsi_score": raw[columns["rsi_score"]],
        "wma_30_score": raw[columns["wma_30_score"]],
        "weekly_rp_score": raw[columns["weekly_rp_score"]],
        "daily_rp_score": raw[columns["daily_rp_score"]],
        "total": raw[columns["total"]],
    })
    df["index_name"] = df["index_name"].map(_norm)
    df["daily_rp_call"] = df["daily_rp_call"].map(_clean_call)
    df["weekly_rp_call"] = df["weekly_rp_call"].map(_clean_call)
    df["wma_30"] = df["wma_30"].map(lambda x: _norm(x).title())
    for col in ["ltp", "rsi", "rsi_score", "wma_30_score", "weekly_rp_score", "daily_rp_score", "total"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df = df[(df["index_name"] != "") & df["total"].notna()].copy()
    if df.empty:
        raise ValueError("Index sheet does not contain any valid index rows.")
    return df.sort_values(["total", "index_name"], ascending=[False, True], kind="stable").reset_index(drop=True)


def read_combined_stock_dataframe(file_bytes: bytes) -> pd.DataFrame:
    raw = _read_stock_sheet_with_detected_header(file_bytes)
    try:
        columns = _require_columns(raw, COMBINED_STOCK_COLS, "stock")
    except ValueError as exc:
        raise ValueError(
            "Could not detect stock file header row. Expected columns like Company Name, Index_Clean, FS, TS, and Total."
        ) from exc
    df = pd.DataFrame({
        "index_clean": raw[columns["index_clean"]],
        "company_name": raw[columns["company_name"]],
        "fs": raw[columns["fs"]],
        "ts": raw[columns["ts"]],
        "total": raw[columns["total"]],
    })
    df["index_clean"] = df["index_clean"].map(_norm)
    df["company_name"] = df["company_name"].map(_norm)
    for col in ["fs", "ts", "total"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df = df[(df["index_clean"] != "") & (df["company_name"] != "") & df["total"].notna()].copy()
    if df.empty:
        raise ValueError("Stock sheet does not contain any valid stock rows.")
    return df.sort_values(["index_clean", "total", "company_name"], ascending=[True, False, True], kind="stable")


def _para(value: Any, style: ParagraphStyle = ST_TBLCELL) -> Paragraph:
    return Paragraph(escape(str(value if value is not None else "")), style)


def _fmt_compact(value: Any, decimals: int = 1) -> str:
    if _isna(value):
        return "-"
    try:
        number = float(value)
    except (TypeError, ValueError):
        return str(value)
    if number.is_integer():
        return str(int(number))
    return f"{number:.{decimals}f}"


def _compact_table(data: list[list[Any]], col_widths: list[float], header_rows: int = 1) -> Table:
    table = Table(data, colWidths=col_widths, repeatRows=header_rows)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, header_rows - 1), CN2),
        ("BACKGROUND", (0, header_rows), (-1, -1), CBG),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, CBDR),
        ("BOX", (0, 0), (-1, -1), 0.5, CBDR),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    return table


def _combined_title(date_str: str) -> list:
    title = Table(
        [[Paragraph("ULJK Bloomberg Ranking System", ST_SMALLTITLE)],
         [Paragraph("Combined Index &amp; Stock Daily Report", ST_CVRSUB)],
         [Paragraph(f"Date: {escape(date_str)}", ST_COVBODY)]],
        colWidths=[USABLE_W],
    )
    title.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), CN),
        ("BOX", (0, 0), (-1, -1), 1.2, CGOLD),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
    ]))
    return [title, Spacer(1, 5 * mm)]


def _index_summary_table(index_df: pd.DataFrame) -> Table:
    headers = ["Rank", "Index Name", "LTP", "Daily", "Weekly", "RSI", "30WMA", "RSI S", "30WMA S", "W RP S", "D RP S", "Total"]
    rows: list[list[Any]] = [[_para(h, ST_TBLHDR) for h in headers]]
    for rank, (_, row) in enumerate(index_df.iterrows(), start=1):
        rows.append([
            _para(rank, ST_TBLNUM),
            _para(row["index_name"]),
            _para(_fmt_compact(row["ltp"], 2), ST_TBLNUM),
            _para(row["daily_rp_call"]),
            _para(row["weekly_rp_call"]),
            _para(_fmt_compact(row["rsi"], 1), ST_TBLNUM),
            _para(row["wma_30"]),
            _para(_fmt_compact(row["rsi_score"], 0), ST_TBLNUM),
            _para(_fmt_compact(row["wma_30_score"], 0), ST_TBLNUM),
            _para(_fmt_compact(row["weekly_rp_score"], 0), ST_TBLNUM),
            _para(_fmt_compact(row["daily_rp_score"], 0), ST_TBLNUM),
            _para(_fmt_compact(row["total"], 0), ST_TBLNUM),
        ])
    widths = [USABLE_W * f for f in [0.045, 0.215, 0.085, 0.07, 0.075, 0.06, 0.065, 0.06, 0.075, 0.075, 0.075, 0.06]]
    return _compact_table(rows, widths)


def _stock_side_by_side_table(top3: pd.DataFrame, bottom3: pd.DataFrame) -> Table:
    rows: list[list[Any]] = [[
        _para("Rank", ST_TBLHDR), _para("Top 3 Stocks", ST_TBLHDR), _para("FS", ST_TBLHDR), _para("TS", ST_TBLHDR), _para("Total", ST_TBLHDR),
        _para("Rank", ST_TBLHDR), _para("Bottom 3 Stocks", ST_TBLHDR), _para("FS", ST_TBLHDR), _para("TS", ST_TBLHDR), _para("Total", ST_TBLHDR),
    ]]
    for i in range(3):
        t = top3.iloc[i] if i < len(top3) else {}
        b = bottom3.iloc[i] if i < len(bottom3) else {}
        rows.append([
            _para(i + 1, ST_TBLNUM),
            _para(t.get("company_name", "")),
            _para(_fmt_compact(t.get("fs"), 1), ST_TBLNUM),
            _para(_fmt_compact(t.get("ts"), 1), ST_TBLNUM),
            _para(_fmt_compact(t.get("total"), 1), ST_TBLNUM),
            _para(i + 1, ST_TBLNUM),
            _para(b.get("company_name", "")),
            _para(_fmt_compact(b.get("fs"), 1), ST_TBLNUM),
            _para(_fmt_compact(b.get("ts"), 1), ST_TBLNUM),
            _para(_fmt_compact(b.get("total"), 1), ST_TBLNUM),
        ])
    widths = [USABLE_W * f for f in [0.05, 0.225, 0.055, 0.055, 0.065, 0.05, 0.225, 0.055, 0.055, 0.065]]
    return _compact_table(rows, widths)


def generate_combined_pdf(stock_file_bytes: bytes, index_file_bytes: bytes, date_str: str) -> bytes:
    index_df = read_combined_index_dataframe(index_file_bytes)
    stock_df = read_combined_stock_dataframe(stock_file_bytes)
    buf = io.BytesIO()
    doc = BaseDocTemplate(buf, pagesize=A4, leftMargin=MARGIN, rightMargin=MARGIN, topMargin=MARGIN, bottomMargin=FOOTER_H + 4 * mm)
    frame = Frame(MARGIN, FOOTER_H + 3 * mm, USABLE_W, PAGE_H - MARGIN - FOOTER_H - 5 * mm, id="main")
    doc.addPageTemplates([PageTemplate(id="combined", frames=[frame], onPage=lambda c, d: _add_footer(c, d, date_str))])
    story: list = []
    story.extend(_combined_title(date_str))
    story.append(_section_banner("PERFORMING INDICES", CGNA))
    story.append(Spacer(1, 2 * mm))
    story.append(_index_summary_table(index_df.head(8)))
    story.append(Spacer(1, 4 * mm))
    story.append(_section_banner("WEAK INDICES", CRDA))
    story.append(Spacer(1, 2 * mm))
    weak = index_df.sort_values(["total", "index_name"], ascending=[True, True], kind="stable").head(8)
    story.append(_index_summary_table(weak))
    story.append(PageBreak())
    for idx_name in sorted(stock_df["index_clean"].unique()):
        sector_df = stock_df[stock_df["index_clean"] == idx_name]
        top3 = sector_df.sort_values(["total", "company_name"], ascending=[False, True], kind="stable").head(3)
        bottom3 = sector_df.sort_values(["total", "company_name"], ascending=[True, True], kind="stable").head(3)
        story.append(_index_banner(idx_name))
        story.append(Spacer(1, 2 * mm))
        story.append(_stock_side_by_side_table(top3, bottom3))
        story.append(Spacer(1, 5 * mm))
    doc.build(story)
    return buf.getvalue()


def _cover_page(date_str: str, indices: list[str]) -> list:
    W = USABLE_W
    elems = []
    elems.append(Spacer(1, 25 * mm))

    title_tbl = Table(
        [[Paragraph("Daily Stock Scoring Report", ST_CVRTITLE)],
         [Paragraph(f"Reporting Date: {date_str}", ST_CVRSUB)]],
        colWidths=[W],
    )
    title_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, -1), CN),
        ("TOPPADDING",    (0, 0), (0, 0),   22),
        ("BOTTOMPADDING", (0, 0), (0, 0),   8),
        ("TOPPADDING",    (0, 1), (0, 1),   4),
        ("BOTTOMPADDING", (0, 1), (0, 1),   22),
        ("BOX",           (0, 0), (-1, -1), 1.5, CGOLD),
    ]))
    elems.append(title_tbl)
    elems.append(Spacer(1, 12 * mm))

    # Indices covered
    idx_data = [[Paragraph(f"•  {idx}", ST_COVBODY)] for idx in sorted(indices)]
    idx_data.insert(0, [Paragraph("Indices Covered", ST_SECHDR)])
    idx_tbl = Table(idx_data, colWidths=[W])
    idx_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (0, 0),   CN2),
        ("BACKGROUND",    (0, 1), (-1, -1), CBG),
        ("TOPPADDING",    (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING",   (0, 0), (-1, -1), 10),
        ("LINEBELOW",     (0, 0), (0, 0),   0.5, CGOLD),
        ("BOX",           (0, 0), (-1, -1), 0.5, CBDR),
    ]))
    elems.append(idx_tbl)
    elems.append(Spacer(1, 10 * mm))

    # Scoring legend
    legend_rows = [
        ("FS — Fundamental Score", "Binary sum of: Sales YoY, Sales QoQ, ROCE, ROE, D/E Ratio, FCF, EPS, ROCE 5Y CAGR, PAT Margin, EBITDA Margin."),
        ("TS — Technical Score",   "Binary sum of: RSI, 30 WMA, Weekly RS (52W), Daily RS (52W)."),
        ("Total Score",            "FS + TS  —  primary ranking column. Higher = stronger stock."),
        ("✓  Tick",                "Indicator score > 0  (positive / bullish signal)."),
        ("✗  Cross",               "Indicator score ≤ 0  (negative / bearish or neutral signal)."),
        ("Top 3",                  "Three highest Total Score stocks in the index  (green cards)."),
        ("Bottom 3",               "Three lowest Total Score stocks in the index  (red cards)."),
    ]

    leg_data = [[Paragraph(k, ST_LEGKEY), Paragraph(v, ST_LEGEND)] for k, v in legend_rows]
    leg_data.insert(0, [Paragraph("Scoring Legend & Indicator Definitions", ST_SECHDR), Paragraph("", ST_LEGEND)])
    leg_tbl = Table(leg_data, colWidths=[W * 0.34, W * 0.66])
    leg_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0),  CN2),
        ("BACKGROUND",    (0, 1), (0, -1),  colors.HexColor("#0d1e30")),
        ("BACKGROUND",    (1, 1), (1, -1),  CBG),
        ("SPAN",          (0, 0), (1, 0)),
        ("TOPPADDING",    (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING",   (0, 0), (-1, -1), 8),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
        ("INNERGRID",     (0, 0), (-1, -1), 0.3, CBDR),
        ("BOX",           (0, 0), (-1, -1), 0.5, CBDR),
        ("VALIGN",        (0, 0), (-1, -1), "TOP"),
        ("LINEBELOW",     (0, 0), (-1, 0),  0.5, CGOLD),
    ]))
    elems.append(leg_tbl)
    elems.append(PageBreak())
    return elems


# ── Main entry point ──────────────────────────────────────────────────────────
def generate_pdf(file_bytes: bytes, date_str: str) -> bytes:
    """Parse the Excel file and return a styled PDF report as bytes."""
    df = read_report_dataframe(file_bytes)
    indices = sorted(df["index_clean"].unique())

    buf = io.BytesIO()

    doc = BaseDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=MARGIN,
        rightMargin=MARGIN,
        topMargin=MARGIN,
        bottomMargin=FOOTER_H + 4 * mm,
    )

    frame = Frame(
        MARGIN,
        FOOTER_H + 3 * mm,
        USABLE_W,
        PAGE_H - MARGIN - FOOTER_H - 5 * mm,
        id="main",
    )

    _date = date_str
    template = PageTemplate(
        id="main_template",
        frames=[frame],
        onPage=lambda c, d: _add_footer(c, d, _date),
    )
    doc.addPageTemplates([template])

    story: list = []
    story.extend(_cover_page(date_str, list(indices)))

    for idx_name in indices:
        idx_df = df[df["index_clean"] == idx_name].copy()
        total_in_idx = len(idx_df)

        top3 = idx_df.nsmallest(3, "top_rank")
        bot3 = idx_df.nsmallest(3, "bot_rank")

        story.append(_index_banner(idx_name))
        story.append(Spacer(1, 4 * mm))

        story.append(_section_banner("▲   TOP 3 STOCKS", CGNA))
        story.append(Spacer(1, 3 * mm))

        for _, row in top3.iterrows():
            story.append(_stock_card(row.to_dict(), int(row["top_rank"]), True, idx_name, total_in_idx))

        story.append(Spacer(1, 5 * mm))
        story.append(_section_banner("▼   BOTTOM 3 STOCKS", CRDA))
        story.append(Spacer(1, 3 * mm))

        for _, row in bot3.iterrows():
            story.append(_stock_card(row.to_dict(), int(row["bot_rank"]), False, idx_name, total_in_idx))

        story.append(PageBreak())

    doc.build(story)
    return buf.getvalue()
