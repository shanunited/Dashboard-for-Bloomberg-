from __future__ import annotations

import argparse
import sys
from datetime import date, datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

DEFAULT_WORKBOOK_PATH = ""

SNAPSHOT_COLUMNS = [
    "Snapshot Date",
    "Index Name",
    "Rank",
    "Stock Name",
    "CMP",
    "Total Score",
    "Technical Score",
    "Fundamental Score",
    "RSI",
    "30WMA",
    "ROCE",
    "ROE",
    "EPS Growth",
    "D/E Ratio",
]


def _clean_header(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def build_headers(workbook_path: Path, sheet_name: str | None = None) -> list[str]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    worksheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]

    row_1 = list(next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True)))
    row_2 = list(next(worksheet.iter_rows(min_row=2, max_row=2, values_only=True)))

    headers: list[str] = []
    seen: dict[str, int] = {}
    active_group = ""

    for top, bottom in zip(row_1, row_2):
        top_text = _clean_header(top)
        bottom_text = _clean_header(bottom)

        if top_text:
            active_group = top_text

        if bottom_text.startswith("FY") and active_group:
            header = f"{active_group} {bottom_text}"
        elif bottom_text:
            header = bottom_text
        elif top_text:
            header = top_text
        else:
            header = "Unnamed"

        count = seen.get(header, 0)
        seen[header] = count + 1
        if count:
            header = f"{header}.{count}"

        headers.append(header)

    workbook.close()
    return headers


def load_metadata(workbook_path: Path, sheet_name: str | None = None) -> pd.DataFrame:
    headers = build_headers(workbook_path, sheet_name=sheet_name)
    dataframe = pd.read_excel(
        workbook_path,
        sheet_name=sheet_name or 0,
        header=None,
        skiprows=2,
        names=headers,
    )
    dataframe = dataframe.dropna(how="all").reset_index(drop=True)
    return dataframe


def pick_column_name(dataframe: pd.DataFrame, candidates: list[str], label: str) -> str:
    for candidate in candidates:
        if candidate in dataframe.columns:
            return candidate
    raise KeyError(f"Missing required column for {label}: {', '.join(candidates)}")


def build_snapshot(
    dataframe: pd.DataFrame,
    snapshot_date: date,
    top_n: int = 10,
) -> pd.DataFrame:
    column_map = {
        "Index Name": pick_column_name(dataframe, ["Index_Clean"], "Index"),
        "Stock Name": pick_column_name(dataframe, ["Company Name"], "Stock Name"),
        "CMP": pick_column_name(dataframe, ["CMP"], "CMP"),
        "Fundamental Score": pick_column_name(dataframe, ["FS"], "Fundamental Score"),
        "Technical Score": pick_column_name(dataframe, ["TS"], "Technical Score"),
        "Total Score": pick_column_name(dataframe, ["Total"], "Total Score"),
        "RSI": pick_column_name(dataframe, ["RSI"], "RSI"),
        "30WMA": pick_column_name(dataframe, ["30WMA", "30 WMA"], "30WMA"),
        "ROCE": pick_column_name(dataframe, ["ROCE FY26", "ROCE FY25"], "ROCE"),
        "ROE": pick_column_name(dataframe, ["ROE FY26", "ROE FY25"], "ROE"),
        "EPS Growth": pick_column_name(dataframe, ["EPS 5Y CAGR"], "EPS Growth"),
        "D/E Ratio": pick_column_name(dataframe, ["D/E"], "D/E Ratio"),
    }

    working = pd.DataFrame(
        {target: dataframe[source] for target, source in column_map.items()}
    )

    numeric_columns = [
        "CMP",
        "Fundamental Score",
        "Technical Score",
        "Total Score",
        "RSI",
        "30WMA",
        "ROCE",
        "ROE",
        "EPS Growth",
        "D/E Ratio",
    ]
    for column in numeric_columns:
        working[column] = pd.to_numeric(working[column], errors="coerce")

    working = working.dropna(subset=["Index Name", "Stock Name", "Total Score"]).copy()
    working["Index Name"] = working["Index Name"].astype(str).str.strip()
    working["Stock Name"] = working["Stock Name"].astype(str).str.strip()

    working = working.sort_values(
        by=["Index Name", "Total Score", "Stock Name"],
        ascending=[True, False, True],
        kind="stable",
    )
    working["Rank"] = working.groupby("Index Name").cumcount() + 1
    working = working[working["Rank"] <= top_n].copy()
    working.insert(0, "Snapshot Date", snapshot_date.strftime("%Y-%m-%d"))

    snapshot = working[SNAPSHOT_COLUMNS].reset_index(drop=True)
    return snapshot


def save_snapshot(snapshot: pd.DataFrame, output_root: Path, snapshot_date: date) -> Path:
    snapshots_dir = output_root / "Daily_Snapshots"
    reports_dir = output_root / "Reports"
    snapshots_dir.mkdir(parents=True, exist_ok=True)
    reports_dir.mkdir(parents=True, exist_ok=True)

    snapshot_path = snapshots_dir / f"{snapshot_date.strftime('%Y_%m_%d')}.csv"
    snapshot.to_csv(snapshot_path, index=False)
    return snapshot_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build a daily Top 10 snapshot from the uploaded Bloomberg metadata workbook."
    )
    parser.add_argument(
        "--input",
        default=DEFAULT_WORKBOOK_PATH,
        help="Path to the uploaded metadata workbook. Leave blank in code and pass it at runtime on the Bloomberg system.",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Optional sheet name. Defaults to the first sheet.",
    )
    parser.add_argument(
        "--output-root",
        default="Bloomberg_Data",
        help="Output folder that will contain Daily_Snapshots and Reports.",
    )
    parser.add_argument(
        "--date",
        default=None,
        help="Snapshot date in YYYY-MM-DD format. Defaults to today.",
    )
    parser.add_argument(
        "--top-n",
        type=int,
        default=10,
        help="Number of ranked rows to keep per index.",
    )
    return parser.parse_args()


def resolve_snapshot_date(raw_value: str | None) -> date:
    if not raw_value:
        return date.today()
    return datetime.strptime(raw_value, "%Y-%m-%d").date()


def main() -> int:
    args = parse_args()
    output_root = Path(args.output_root).expanduser().resolve()

    if not args.input:
        print(
            "Workbook path is empty. Pass --input on the Bloomberg system, for example "
            '--input "C:\\Users\\Shantanu\\Desktop\\Shantanu\\Daily_File\\Bloomberg_Daily_04_22_2026.xlsx"',
            file=sys.stderr,
        )
        return 1

    workbook_path = Path(args.input).expanduser().resolve()

    if not workbook_path.exists():
        print(f"Workbook not found: {workbook_path}", file=sys.stderr)
        return 1

    try:
        snapshot_date = resolve_snapshot_date(args.date)
        metadata = load_metadata(workbook_path, sheet_name=args.sheet)
        snapshot = build_snapshot(metadata, snapshot_date=snapshot_date, top_n=args.top_n)
        output_path = save_snapshot(snapshot, output_root=output_root, snapshot_date=snapshot_date)
    except Exception as exc:  # pragma: no cover - CLI safety path
        print(f"Snapshot generation failed: {exc}", file=sys.stderr)
        return 1

    print(f"Snapshot created: {output_path}")
    print(f"Rows written: {len(snapshot)}")
    print(f"Indices covered: {snapshot['Index Name'].nunique()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
